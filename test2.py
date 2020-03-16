from openpyxl import load_workbook
wb = load_workbook('F:\data2.xlsx')
ws = wb.active
n = 100000  #总人数
i = 100  #感染人数
s = n -i  #未感染人数/易感染人数
a = 0.03  #感染率
b = 0.5  #恢复率
times = 0
while(times>100000 or i >= n or n <=0):
    it = s*a*i/n
    r = i*b
    i = it + i - r
    times=times + 1
    ws.cell(row=times,column=1).value=i
    wb.save('F:\data2.xlsx')
    print i
