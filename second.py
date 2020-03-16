# -*- coding: cp936 -*-
from openpyxl import load_workbook
wb = load_workbook('F:\data.xlsx')
ws = wb.active
r = [2.2,2.3,2.4,2.5,2.6,2.7,2.8,2.9,3.0,3.1,3.2,3.3,3.4,3.5,3.6,3.7,3.8,3.9,4,4.1,4.2,4.3,4.4,4.5,4.6,4.7,4.8,4.9,5,5.1,5.2,5.3,5.4,5.5,5.6,5.7,5.8,5.9,6,6.1,6.2,6.3,6.4,6.5,6.6,6.7,6.8,6.9,7,7.1,7.2,7.3,7.4,7.5,7.6,7.7,7.8,7.9,8,]  #2.2-3.6
e = [0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]
v = []
for i in range(0,len(r)):
    rt = r[i]
    for j in range(0,len(e)):
        et = e[j]
        vt = (1-1/rt)/et
        v.append(vt)
        ws.cell(row=i+1,column=j+1).value=vt
        wb.save('F:/data.xlsx')
        print vt
    print('*****')
print('结果列表长度：{}'.format(len(v)))
print('结果最小值：{}'.format(min(v)))
print('结果最大值：{}'.format(max(v)))

for k in range(0,len(v)):
    if v[k] == min(v):
        print('最小值位置：{}'.format(k))
    elif v[k] == max(v):
        print('最大值位置：{}'.format(k))


