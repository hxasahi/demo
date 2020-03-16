import random
from openpyxl import load_workbook
wb = load_workbook('F:\data3.xlsx')
ws = wb.active
'''def gx():
    #m = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    m = [1,2,3,4,5]
    return random.choice(m)'''

N = 10000
I = 5
S = N - I

e = S/N
#J = 300
J = 60

X = 10000

a = 1
b = 0.5
c = 0.05
d = 0.001
f = 0

z = I*f
r = z*b
m = I*c
K = I*d

times = 0

while (times < 10000 and I > 0):
    if(times == 10):
        #J = 100
        #J = 30
        f = 0.3
    elif(times == 20):
        #J = 10
        f = 0.5
    elif(times == 30):
        #J = 5
        f = 1
    flag = 0
    times = times + 1
    I = I + m - r - K + J*a
    if I > S-K :
        flag = 1
        break
    z = I*f
    r = r + z*b
    temp = r
    if I > X :
        r = 0
    m = m + I*c
    K = K + I*d
    #J = gx()
    S = N - I
    ws.cell(row=times,column=6).value=I
    
    print int(I)
wb.save('F:\data3.xlsx')
if (flag == 0) :
    print('win')
    print ('recover: {}'.format(temp))
    print ('kill: {}'.format(K))
    print ('times: {}'.format(times))
    print ('people_num: {}'.format(S-K))
elif (flag == 1) :
    print('fale')
    print ('recover: {}'.format(temp))
    print ('kill: {}'.format(K))
    print ('times: {}'.format(times))
    print ('people_num: {}'.format(S-K))
